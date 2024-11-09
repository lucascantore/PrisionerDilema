# players
from players.teammates_lucifer.TeammateLuciferPointGiver import TeammateLuciferPointGiver
from players.teammates_lucifer.TeammateLuciferPointReceiver import TeammateLuciferPointReceiver
from players.Lucifer import Lucifer
from players.Random import Random
from players.GoodBoy import GoodBoy
from players.TipForTat import TipForTat
from players.teammates_tipfortat.TeammateTipForTatPointGiver import TeammateTipForTatPointGiver
from players.teammates_tipfortat.TeammateTipForTatPointReceiver import TeammateTipForTatPointReceiver
from players.teammate_dalecooper.TeammateDaleCooperGiver import TeammateDaleCooperGiver
from players.teammate_dalecooper.TeammateDaleCooperReceiver import TeammateDaleCooperReceiver
from players.students_players.DaleCooper import DaleCooper

# Tools
from experimentationTools import run_experiment_with, list_of_players
from itertools import product
import openpyxl

from experimentation_with_students_players import run_experiment_with_students_players, \
    run_experiment_with_student_players_and_teammate_dale_cooper, test_all_out_war_with_less_participants, \
    test_all_out_war_single_tournament, test_all_out_war_high_numbers, test_all_out_war_high_numbers_and_nathan, \
    test_all_out_war_with_less_randomness


def teammates_lucifer_vs_lucifer(max_rounds):
    run_experiment_with(
        [TeammateLuciferPointGiver, TeammateLuciferPointReceiver]
        + list_of_players(Lucifer, 1),
        max_rounds,
        "experiment_results/1_teammates_lucifer_vs_lucifer.xlsx"
    )


def teammates_lucifer_vs_random_vs_lucifer(max_rounds):
    run_experiment_with(
        [TeammateLuciferPointGiver, TeammateLuciferPointReceiver]
        + list_of_players(Lucifer, 1)
        + list_of_players(Random, 5),
        max_rounds,
        "experiment_results/2_teammates_lucifer_vs_random_vs_lucifer.xlsx"
    )


def teammates_lucifer_vs_goodboy_vs_lucifer(max_rounds):
    run_experiment_with(
        [TeammateLuciferPointGiver, TeammateLuciferPointReceiver]
        + list_of_players(Lucifer, 1)
        + list_of_players(GoodBoy, 5),
        max_rounds,
        "experiment_results/3_teammates_lucifer_vs_goodboy_vs_lucifer.xlsx"
    )


def teammates_lucifer_vs_goodboy_vs_random_vs_lucifer(max_rounds):
    run_experiment_with(
        [TeammateLuciferPointGiver, TeammateLuciferPointReceiver]
        + list_of_players(Lucifer, 5)
        + list_of_players(GoodBoy, 5)
        + list_of_players(Random, 5),
        max_rounds,
        "experiment_results/4_teammates_lucifer_vs_goodboy_vs_random_vs_lucifer.xlsx"
    )


def teammates_lucifer_vs_goodboy_vs_random_vs_single_tipfortat(max_rounds):
    run_experiment_with(
        [TeammateLuciferPointGiver, TeammateLuciferPointReceiver]
        + list_of_players(TipForTat, 1)
        + list_of_players(GoodBoy, 5)
        + list_of_players(Random, 5),
        max_rounds,
        "experiment_results/5_teammates_lucifer_vs_goodboy_vs_random_vs_single_tipfortat.xlsx"
    )


def teammates_lucifer_vs_goodboy_vs_random_vs_ten_tipfortat(max_rounds):
    run_experiment_with(
        [TeammateLuciferPointGiver, TeammateLuciferPointReceiver]
        + list_of_players(TipForTat, 10)
        + list_of_players(GoodBoy, 5)
        + list_of_players(Random, 5),
        max_rounds,
        "experiment_results/6_teammates_lucifer_vs_goodboy_vs_random_vs_ten_tipfortat.xlsx"
    )


def teammates_lucifer_vs_goodboy_vs_random_vs_fifteen_tipfortat(max_rounds):
    run_experiment_with(
        [TeammateLuciferPointGiver, TeammateLuciferPointReceiver]
        + list_of_players(TipForTat, 15)
        + list_of_players(GoodBoy, 5)
        + list_of_players(Random, 5),
        max_rounds,
        "experiment_results/7_teammates_lucifer_vs_goodboy_vs_random_vs_fifteen_tipfortat.xlsx"
    )


def teammates_lucifer_vs_three_tipfortat(max_rounds):
    run_experiment_with(
        [TeammateLuciferPointGiver, TeammateLuciferPointReceiver]
        + list_of_players(TipForTat, 3),
        max_rounds,
        "experiment_results/8_teammates_lucifer_vs_three_tipfortat.xlsx"
    )


def teammates_tipfortat_vs_tipfortat(max_rounds):
    run_experiment_with(
        [TeammateTipForTatPointGiver, TeammateTipForTatPointReceiver]
        + list_of_players(TipForTat, 5),
        max_rounds,
        "experiment_results/9_teammates_tipfortat_vs_tipfortat.xlsx"
    )


def teammates_tipfortat_vs_random_vs_tipfortat(max_rounds):
    run_experiment_with(
        [TeammateTipForTatPointGiver, TeammateTipForTatPointReceiver]
        + list_of_players(TipForTat, 5)
        + list_of_players(Random, 5),
        max_rounds,
        "experiment_results/10_teammates_tipfortat_vs_random_vs_tipfortat.xlsx"
    )


def teammates_tipfortat_vs_random_vs_goodboy_vs_tipfortat(max_rounds):
    run_experiment_with(
        [TeammateTipForTatPointGiver, TeammateTipForTatPointReceiver]
        + list_of_players(TipForTat, 5)
        + list_of_players(Random, 5)
        + list_of_players(GoodBoy, 5),
        max_rounds,
        "experiment_results/11_teammates_tipfortat_vs_random_vs_goodboy_vs_tipfortat.xlsx"
    )


def teammates_tipfortat_vs_random_vs_goodboy_vs_lucifer_vs_tipfortat(max_rounds):
    run_experiment_with(
        [TeammateTipForTatPointGiver, TeammateTipForTatPointReceiver]
        + list_of_players(TipForTat, 5)
        + list_of_players(Random, 5)
        + list_of_players(GoodBoy, 5)
        + list_of_players(Lucifer, 3),
        max_rounds,
        "experiment_results/12_teammates_tipfortat_vs_random_vs_goodboy_vs_lucifer_vs_tipfortat.xlsx"
    )


def teammates_tipfortat_vs_random_vs_goodboy_vs_lucifer(max_rounds):
    run_experiment_with(
        [TeammateTipForTatPointGiver, TeammateTipForTatPointReceiver]
        + list_of_players(Random, 5)
        + list_of_players(GoodBoy, 5)
        + list_of_players(Lucifer, 1),
        max_rounds,
        "experiment_results/13_teammates_tipfortat_vs_random_vs_goodboy_vs_lucifer.xlsx"
    )


def teammates_tipfortat_vs_lucifer(max_rounds):
    run_experiment_with(
        [TeammateTipForTatPointGiver, TeammateTipForTatPointReceiver]
        + list_of_players(Lucifer, 1),
        max_rounds,
        "experiment_results/14_teammates_tipfortat_vs_lucifer.xlsx"
    )


def teammates_tipfortat_vs_random_vs_lucifer(max_rounds):
    run_experiment_with(
        [TeammateTipForTatPointGiver, TeammateTipForTatPointReceiver]
        + list_of_players(Random, 5)
        + list_of_players(Lucifer, 1),
        max_rounds,
        "experiment_results/15_teammates_tipfortat_vs_random_vs_lucifer.xlsx"
    )


def teammates_tipfortat_vs_teammates_lucifer_vs_random_vs_tipfortat_vs_lucifer_vs_goodboy(max_rounds):
    run_experiment_with(
        [TeammateTipForTatPointGiver, TeammateTipForTatPointReceiver,
         TeammateLuciferPointGiver, TeammateLuciferPointReceiver]
        + list_of_players(Random, 5)
        + list_of_players(TipForTat, 5)
        + list_of_players(Lucifer, 5)
        + list_of_players(GoodBoy, 5),
        max_rounds,
        "experiment_results/16_teammates_tipfortat_vs_teammates_lucifer_vs_random_vs_tipfortat_vs_lucifer_vs_goodboy.xlsx"
    )


def multiple_everything_with_all_classic_agents(max_rounds):
    range_of_players = range(1, 6)
    result_number = 19
    for num_random_players, num_tipfortat_players, num_lucifer_players, num_goodboy_players in product(range_of_players,
                                                                                                       repeat=4):
        run_experiment_with(
            [TeammateTipForTatPointGiver, TeammateTipForTatPointReceiver,
             TeammateLuciferPointGiver, TeammateLuciferPointReceiver]
            + list_of_players(Random, num_random_players)
            + list_of_players(TipForTat, num_tipfortat_players)
            + list_of_players(Lucifer, num_lucifer_players)
            + list_of_players(GoodBoy, num_goodboy_players),
            max_rounds,
            f"experiment_results/{result_number}_teammates_tipfortat_vs_teammates_lucifer_vs_{num_random_players}random_vs_{num_tipfortat_players}tipfortat_vs_{num_lucifer_players}lucifer_vs_{num_goodboy_players}goodboy.xlsx"
        )
        result_number = result_number + 1


def analyze_final_results():
    range_of_players = range(1, 6)
    result_number = 19

    # Create a new workbook to store the results
    result_wb = openpyxl.Workbook()
    result_sheet = result_wb.active
    result_sheet.title = "Analysis Results"

    # Write headers to the result sheet
    result_sheet.cell(row=1, column=1, value="File Name")
    result_sheet.cell(row=1, column=2, value="Value from Row 1")

    # Variable to keep track of the row in the result sheet
    result_row = 2

    for num_random_players, num_tipfortat_players, num_lucifer_players, num_goodboy_players in product(range_of_players,
                                                                                                       repeat=4):
        filepath = f"experiment_results/{result_number}_teammates_tipfortat_vs_teammates_lucifer_vs_{num_random_players}random_vs_{num_tipfortat_players}tipfortat_vs_{num_lucifer_players}lucifer_vs_{num_goodboy_players}goodboy.xlsx"

        # Load the workbook and select the active sheet
        wb = openpyxl.load_workbook(filepath)
        sheet = wb.active

        # Find the number of columns between 9 and 21 (in row 50)
        row_50_values = []
        for col in range(2, 22):  # Check up to column 21 (assuming max 21 columns)
            cell_value = sheet.cell(row=50, column=col).value
            if cell_value is not None:  # If there's a value in the cell, consider it
                row_50_values.append(cell_value)

        # Get the column with the highest value in row 50
        max_value = max(row_50_values)
        max_column_index = row_50_values.index(max_value) + 2  # Column index starts from 1

        # Get the value from row 1 in the same column
        value_in_row_1 = sheet.cell(row=1, column=max_column_index).value

        # Write the result to the new Excel sheet
        result_sheet.cell(row=result_row, column=1, value=filepath)  # File name
        result_sheet.cell(row=result_row, column=2, value=value_in_row_1)  # Value from row 1

        # Move to the next row in the result sheet
        result_row += 1

        result_number = result_number + 1

    result_wb.save("experiment_results/analysis_results.xlsx")


def run_all_experiments():
    max_rounds = 1300
    # teammates_lucifer_vs_lucifer(max_rounds)
    # teammates_lucifer_vs_random_vs_lucifer(max_rounds)
    # teammates_lucifer_vs_goodboy_vs_lucifer(max_rounds)
    # teammates_lucifer_vs_goodboy_vs_random_vs_lucifer(max_rounds)
    # teammates_lucifer_vs_goodboy_vs_random_vs_single_tipfortat(max_rounds)
    # teammates_lucifer_vs_goodboy_vs_random_vs_ten_tipfortat(max_rounds)
    # teammates_lucifer_vs_goodboy_vs_random_vs_fifteen_tipfortat(max_rounds)
    # teammates_lucifer_vs_three_tipfortat(max_rounds)
    # teammates_tipfortat_vs_tipfortat(max_rounds)
    # teammates_tipfortat_vs_random_vs_tipfortat(max_rounds)
    # teammates_tipfortat_vs_random_vs_goodboy_vs_tipfortat(max_rounds)
    # teammates_tipfortat_vs_random_vs_goodboy_vs_lucifer_vs_tipfortat(max_rounds)
    # teammates_tipfortat_vs_random_vs_goodboy_vs_lucifer(max_rounds)
    # teammates_tipfortat_vs_lucifer(max_rounds)
    # teammates_tipfortat_vs_random_vs_lucifer(max_rounds)
    # teammates_tipfortat_vs_teammates_lucifer_vs_random_vs_tipfortat_vs_lucifer_vs_goodboy(max_rounds)

    # run_experiment_with_students_players(max_rounds)
    # run_experiment_with_student_players_and_teammate_dale_cooper(max_rounds)

    # test_all_out_war_with_less_participants(max_rounds)
    # test_all_out_war_single_tournament(max_rounds)
    # test_all_out_war_high_numbers(200, max_rounds)
    # test_all_out_war_high_numbers_and_nathan(1100, 1300) 
    # test_all_out_war_with_less_randomness(5, 2900)

    # multiple_everything_with_all_classic_agents(max_rounds)
    # analyze_final_results()
