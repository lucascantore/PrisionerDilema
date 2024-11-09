from Prisoner import create_dynamic_prisoner_class

# tournament
from Tournament import run_round_robin_tournament

# excel formatting
import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill
import math


def list_of_players(base_player, amount_of_players):
    player_list = []
    for i in range(amount_of_players):
        player_list = player_list + [create_dynamic_prisoner_class(base_player, f"{base_player.name()}_{i}")]

    return player_list


# we only support up to 675 players
def excel_column_number_for(number):
    first_letter_index = math.floor(number / 26)
    second_letter_index = number % 26

    if first_letter_index > 0:
        return f"{chr(65 + first_letter_index)}{chr(65 + second_letter_index)}"
    else:
        return f"{chr(65 + second_letter_index)}"


def format_results_to_excel(results, competing_players, excel_name):
    data = {
        "Games Played": list(results.keys())
    }

    for element in competing_players:
        data[element.name()] = []

    for key in results:
        for element in results[key]:
            data[element["name"]] = data[element["name"]] + [element["score"]]

    df = pd.DataFrame(data)
    df.to_excel(excel_name, index=False)

    wb = load_workbook(excel_name)
    ws = wb.active

    highlight_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    for row in range(2, len(results.keys()) + 2):
        formula = f'B{row}=MAX($B{row}:${excel_column_number_for(len(competing_players))}{row})'
        rule = FormulaRule(formula=[formula], fill=highlight_fill)

        # Apply the rule to the row
        ws.conditional_formatting.add(f'B{row}:{excel_column_number_for(len(competing_players))}{row}', rule)

    wb.save(excel_name)  # Save to a new file or overwrite the existing one


def run_experiment_with(competing_agents, max_n_rounds, excel_name):
    print("started experiment for: ", list(map(lambda a: a.name(), competing_agents)))
    results = {}
    n_rounds = 5
    while n_rounds <= max_n_rounds:
        print("start round numer: ", n_rounds)
        results[f"{n_rounds}"] = run_round_robin_tournament(competing_agents, n_rounds)
        if n_rounds < 100:
            n_rounds += 5
        else:
            n_rounds += 10

    format_results_to_excel(results, competing_agents, excel_name)


def run_single_experiment_with(competing_agents, n_rounds, excel_name):
    print("started experiment for: ", list(map(lambda a: a.name(), competing_agents)))
    results = {f"{n_rounds}": run_round_robin_tournament(competing_agents, n_rounds)}
    format_results_to_excel(results, competing_agents, excel_name)

def run_experiment_with_between(competing_agents, n_rounds_start, n_rounds_finish, excel_name):
    print("started experiment for: ", list(map(lambda a: a.name(), competing_agents)))
    results = {}
    n_rounds = n_rounds_start
    while n_rounds <= n_rounds_finish:
        print("start round numer: ", n_rounds)
        results[f"{n_rounds}"] = run_round_robin_tournament(competing_agents, n_rounds)
        n_rounds += 100

    format_results_to_excel(results, competing_agents, excel_name)